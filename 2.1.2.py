import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side
import matplotlib.pyplot as plt
import numpy as np


class Reps:
    def __init__(self, vacs_n, stat_f, stat_s, stat_t, stat_four, stat_five, stat_six):
        self.wb = Workbook()
        self.vacancy_name = vacs_n
        self.stats1 = stat_f
        self.stats2 = stat_s
        self.stats3 = stat_t
        self.stats4 = stat_four
        self.stats5 = stat_five
        self.stats6 = stat_six

    def exce(self):
        w_s1 = self.wb.active
        w_s1.title = 'Статистика по годам'
        w_s1.append(['Год', 'Средняя зарплата', 'Средняя зарплата - ' + self.vacancy_name, 'Количество вакансий',
                     'Количество вакансий - ' + self.vacancy_name])
        for y in self.stats1.keys():
            w_s1.append([y, self.stats1[y], self.stats3[y], self.stats2[y], self.stats4[y]])

        datas = [['Год ', 'Средняя зарплата ', ' Средняя зарплата - ' + self.vacancy_name, ' Количество вакансий',
                  ' Количество вакансий - ' + self.vacancy_name]]
        cols_w = []
        for line in datas:
            for x, z in enumerate(line):
                if len(cols_w) > x:
                    if len(z) > cols_w[x]:
                        cols_w[x] = len(z)
                else:
                    cols_w += [len(z)]

        for x, cols_w in enumerate(cols_w, 1):
            w_s1.column_dimensions[get_column_letter(x)].width = cols_w + 2

        datas = []
        datas.append(['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий'])
        for (city_f, v_f), (city_s, v_s) in zip(self.stats5.items(), self.stats6.items()):
            datas.append([city_f, v_f, '', city_s, v_s])
        w_s2 = self.wb.create_sheet('Статистика по городам')
        for line in datas:
            w_s2.append(line)

        cols_w = []
        for line in datas:
            for x, z in enumerate(line):
                z = str(z)
                if len(cols_w) > x:
                    if len(z) > cols_w[x]:
                        cols_w[x] = len(z)
                else:
                    cols_w += [len(z)]

        for x, cols_w in enumerate(cols_w, 1):
            w_s2.column_dimensions[get_column_letter(x)].width = cols_w + 2

        b_fonts = Font(bold=True)
        for c in 'ABCDE':
            w_s1[c + '1'].font = b_fonts
            w_s2[c + '1'].font = b_fonts

        for i, _ in enumerate(self.stats5):
            w_s2['E' + str(i + 2)].number_format = '0.00%'

        th = Side(border_style='thin',
                  color='00000000')

        for line in range(len(datas)):
            for c in 'ABDE':
                w_s2[c + str(line + 1)].border = Border(left=th,
                                                        bottom=th,
                                                        right=th,
                                                        top=th)

        for line, _ in enumerate(self.stats1):
            for c in 'ABCDE':
                w_s1[c + str(line + 1)].border = Border(left=th,
                                                        bottom=th,
                                                        right=th,
                                                        top=th)

    def img(self):
        figs, ((x_f, x_s), (x_t, x_four)) = plt.subplots(nrows=2,
                                                         ncols=2)

        b_f = x_f.bar(np.array(list(self.stats1.keys())) - 0.4, self.stats1.values(), width=0.4)
        b_s = x_f.bar(np.array(list(self.stats1.keys())), self.stats3.values(), width=0.4)
        x_f.set_title('Уровень зарплат по годам', fontdict={'fontsize': 8})
        x_f.grid(axis='y')
        x_f.legend((b_f[0], b_s[0]), ('средняя з/п', 'з/п ' + self.vacancy_name.lower()), prop={'size': 8})
        x_f.set_xticks(np.array(list(self.stats1.keys())) - 0.2, list(self.stats1.keys()), rotation=90)
        x_f.xaxis.set_tick_params(labelsize=8)
        x_f.yaxis.set_tick_params(labelsize=8)

        x_s.set_title('Количество вакансий по годам', fontdict={'fontsize': 8})
        b_f = x_s.bar(np.array(list(self.stats2.keys())) - 0.4, self.stats2.values(), width=0.4)
        b_s = x_s.bar(np.array(list(self.stats2.keys())), self.stats4.values(), width=0.4)
        x_s.legend((b_f[0], b_s[0]), ('Количество вакансий', 'Количество вакансий\n' + self.vacancy_name.lower()),
                   prop={'size': 8})
        x_s.set_xticks(np.array(list(self.stats2.keys())) - 0.2, list(self.stats2.keys()), rotation=90)
        x_s.grid(axis='y')
        x_s.xaxis.set_tick_params(labelsize=8)
        x_s.yaxis.set_tick_params(labelsize=8)

        x_t.set_title('Уровень зарплат по городам', fontdict={'fontsize': 8})
        x_t.barh(list([str(a).replace(' ', '\n').replace('-', '-\n') for a in reversed(list(self.stats5.keys()))]),
                 list(reversed(list(self.stats5.values()))), color='blue', height=0.5, align='center')
        x_t.yaxis.set_tick_params(labelsize=6)
        x_t.xaxis.set_tick_params(labelsize=8)
        x_t.grid(axis='x')

        x_four.set_title('Доля вакансий по городам', fontdict={'fontsize': 8})
        other = 1 - sum([value for value in self.stats6.values()])
        x_four.pie(list(self.stats6.values()) + [other], labels=list(self.stats6.keys()) + ['Другие'],
                   textprops={'fontsize': 6})

        plt.tight_layout()
        plt.savefig('graph.png')

    def save(self, name):
        self.wb.save(filename=name)


class Inputs:
    def __init__(self):
        self.file_name = input('Введите название файла: ')
        self.vacancy_name = input('Введите название профессии: ')

        setsdatas = DatasSets(self.file_name, self.vacancy_name)
        stat_f, stat_s, stat_t, stat_four, stat_five, stat_six = setsdatas.stats()
        setsdatas.outs(stat_f, stat_s, stat_t, stat_four, stat_five, stat_six)

        reps = Reps(self.vacancy_name,
                    stat_f,
                    stat_s,
                    stat_t,
                    stat_four,
                    stat_five,
                    stat_six)
        reps.exce()
        reps.save('report.xlsx')
        reps.img()

class DatasSets:
    def __init__(self, names_f, vacs_n):
        self.file_name = names_f
        self.vacancy_name = vacs_n

    @staticmethod
    def cram(dic, k, a):
        if k in dic:
            dic[k] += a
        else:
            dic[k] = a

    @staticmethod
    def avers(dic):
        dic_n = {}
        for k, v in dic.items():
            dic_n[k] = int(sum(v) / len(v))
        return dic_n

    def readers(self):
        with open(self.file_name, mode='r', encoding='utf-8-sig') as f:
            r = csv.reader(f)
            h = next(r)
            lens = len(h)
            for line in r:
                if '' not in line and len(line) == lens:
                    yield dict(zip(h, line))

    def stats(self):
        sal = {}
        sal_vn = {}
        sal_cities = {}
        count_vacs = 0

        for vac_dic in self.readers():
            vacs = Vacs(vac_dic)
            self.cram(sal, vacs.year, [vacs.salary_average])
            if vacs.name.find(self.vacancy_name) != -1:
                self.cram(sal_vn, vacs.year, [vacs.salary_average])
            self.cram(sal_cities, vacs.area_name, [vacs.salary_average])
            count_vacs += 1

        vac_num = dict([(k, len(v)) for k, v in sal.items()])
        vac_num_n = dict([(k, len(v)) for k, v in sal_vn.items()])

        if not sal_vn:
            sal_vn = dict([(k, [0]) for k, v in sal.items()])
            vac_num_n = dict([(k, 0) for k, v in vac_num.items()])

        stat_f = self.avers(sal)
        stat_s = self.avers(sal_vn)
        stat_t = self.avers(sal_cities)

        stat_four = {}
        for y, sals in sal_cities.items():
            stat_four[y] = round(len(sals) / count_vacs, 4)
        stat_four = list(filter(lambda s: s[-1] >= 0.01, [(k, v) for k, v in stat_four.items()]))
        stat_four.sort(key=lambda s: s[-1],
                       reverse=True)
        stat_five = stat_four.copy()
        stat_four = dict(stat_four)
        stat_t = list(filter(lambda s: s[0] in list(stat_four.keys()), [(k, v) for k, v in stat_t.items()]))
        stat_t.sort(key=lambda s: s[-1],
                    reverse=True)
        stat_t = dict(stat_t[:10])
        stat_five = dict(stat_five[:10])

        return stat_f, vac_num, stat_s, vac_num_n, stat_t, stat_five

    @staticmethod
    def outs(stat_f, stat_s, stat_t, stat_four, stat_five, stat_six):
        print('Динамика уровня зарплат по годам: {0}'.format(stat_f))
        print('Динамика количества вакансий по годам: {0}'.format(stat_s))
        print('Динамика уровня зарплат по годам для выбранной профессии: {0}'.format(stat_t))
        print('Динамика количества вакансий по годам для выбранной профессии: {0}'.format(stat_four))
        print('Уровень зарплат по городам (в порядке убывания): {0}'.format(stat_five))
        print('Доля вакансий по городам (в порядке убывания): {0}'.format(stat_six))

class Vacs:
    currency_to_rub = {
        "AZN": 35.68,
        "BYR": 23.91,
        "EUR": 59.90,
        "GEL": 21.74,
        "KGS": 0.76,
        "KZT": 0.13,
        "RUR": 1,
        "UAH": 1.64,
        "USD": 60.66,
        "UZS": 0.0055
    }

    def __init__(self, vacs):
        self.name = vacs['name']
        self.salary_from = int(float(vacs['salary_from']))
        self.salary_to = int(float(vacs['salary_to']))
        self.salary_currency = vacs['salary_currency']
        self.salary_average = self.currency_to_rub[self.salary_currency] * (self.salary_from + self.salary_to) / 2
        self.area_name = vacs['area_name']
        self.year = int(vacs['published_at'][:4])


if __name__ == '__main__':
    Inputs()