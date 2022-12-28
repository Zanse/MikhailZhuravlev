import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side
import matplotlib.pyplot as plt
import numpy as np
from jinja2 import Environment, FileSystemLoader
import pathlib
import pdfkit

class Inputs:
    def __init__(self):
        self.file_name = input('Введите название файла: ')
        self.vacancy_name = input('Введите название профессии: ')

        datassets = DatasSets(self.file_name, self.vacancy_name)
        stats1, stat_s, stat_t, stat_four, stat_five, stat_six = datassets.g_stats()
        datassets.outs(stats1, stat_s, stat_t, stat_four, stat_five, stat_six)

        reps = Reps(self.vacancy_name, stats1, stat_s, stat_t, stat_four, stat_five, stat_six)
        reps.exce()
        reps.img()
        reps.pdfs()


class DatasSets:
    def readers(self):
        with open(self.file_name, mode='r', encoding='utf-8-sig') as f:
            r = csv.reader(f)
            h = next(r)
            lens = len(h)
            for lines in r:
                if '' not in lines and len(lines) == lens:
                    yield dict(zip(h, lines))

    @staticmethod
    def creams(dics, k, a):
        if k in dics:
            dics[k] += a
        else:
            dics[k] = a

    def __init__(self, names, vacs_n):
        self.file_name = names
        self.vacancy_name = vacs_n

    @staticmethod
    def avers(dics):
        dics_n = {}
        for k, v in dics.items():
            dics_n[k] = int(sum(v) / len(v))
        return dics_n

    def g_stats(self):
        sals = {}
        slas_vac = {}
        sals_cities = {}
        count_v = 0

        for vac_d in self.readers():
            vacs = Vacs(vac_d)
            self.creams(sals, vacs.year, [vacs.salary_average])
            if vacs.name.find(self.vacancy_name) != -1:
                self.creams(slas_vac, vacs.year, [vacs.salary_average])
            self.creams(sals_cities, vacs.area_name, [vacs.salary_average])
            count_v += 1

        vacs_n = dict([(k, len(v)) for k, v in sals.items()])
        vacs_nn = dict([(k, len(v)) for k, v in slas_vac.items()])

        if not slas_vac:
            slas_vac = dict([(k, [0]) for k, v in sals.items()])
            vacs_nn = dict([(k, 0) for k, v in vacs_n.items()])

        stat_f = self.avers(sals)
        stat_s = self.avers(slas_vac)
        stat_t = self.avers(sals_cities)

        stat_four = {}
        for y, sal in sals_cities.items():
            stat_four[y] = round(len(sal) / count_v, 4)
        stat_four = list(filter(lambda s: s[-1] >= 0.01, [(k, v) for k, v in stat_four.items()]))
        stat_four.sort(key=lambda s: s[-1], reverse=True)
        stat_five = stat_four.copy()
        stat_four = dict(stat_four)
        stat_t = list(filter(lambda s: s[0] in list(stat_four.keys()), [(k, v) for k, v in stat_t.items()]))
        stat_t.sort(key=lambda s: s[-1], reverse=True)
        stat_t = dict(stat_t[:10])
        stat_five = dict(stat_five[:10])

        return stat_f, vacs_n, stat_s, vacs_nn, stat_t, stat_five

    @staticmethod
    def outs(stat_f, stat_s, stat_t, stat_four, stat_five, stat_six):
        print('Динамика уровня зарплат по годам: {0}'.format(stat_f))
        print('Динамика количества вакансий по годам: {0}'.format(stat_s))
        print('Динамика уровня зарплат по годам для выбранной профессии: {0}'.format(stat_t))
        print('Динамика количества вакансий по годам для выбранной профессии: {0}'.format(stat_four))
        print('Уровень зарплат по городам (в порядке убывания): {0}'.format(stat_five))
        print('Доля вакансий по городам (в порядке убывания): {0}'.format(stat_six))


class Reps:
    def __init__(self, names, stat_f, stat_s, stat_t, stat_four, stat_five, stat_six):
        self.wb = Workbook()
        self.vacancy_name = names
        self.stats1 = stat_f
        self.stats2 = stat_s
        self.stats3 = stat_t
        self.stats4 = stat_four
        self.stats5 = stat_five
        self.stats6 = stat_six

    def img(self):
        fi, ((a_f, a_s), (a_t, a_four)) = plt.subplots(nrows=2,
                                                     ncols=2)

        b_f = a_f.bar(np.array(list(self.stats1.keys())) - 0.4, self.stats1.values(), width=0.4)
        b_s = a_f.bar(np.array(list(self.stats1.keys())), self.stats3.values(), width=0.4)
        a_f.set_title('Уровень зарплат по годам', fontdict={'fontsize': 8})
        a_f.grid(axis='y')
        a_f.legend((b_f[0], b_s[0]), ('средняя з/п', 'з/п ' + self.vacancy_name.lower()), prop={'size': 8})
        a_f.set_xticks(np.array(list(self.stats1.keys())) - 0.2, list(self.stats1.keys()), rotation=90)
        a_f.xaxis.set_tick_params(labelsize=8)
        a_f.yaxis.set_tick_params(labelsize=8)

        a_s.set_title('Количество вакансий по годам', fontdict={'fontsize': 8})
        b_f = a_s.bar(np.array(list(self.stats2.keys())) - 0.4, self.stats2.values(), width=0.4)
        b_s = a_s.bar(np.array(list(self.stats2.keys())), self.stats4.values(), width=0.4)
        a_s.legend((b_f[0], b_s[0]), ('Количество вакансий', 'Количество вакансий\n' + self.vacancy_name.lower()), prop={'size': 8})
        a_s.set_xticks(np.array(list(self.stats2.keys())) - 0.2, list(self.stats2.keys()), rotation=90)
        a_s.grid(axis='y')
        a_s.xaxis.set_tick_params(labelsize=8)
        a_s.yaxis.set_tick_params(labelsize=8)

        a_t.set_title('Уровень зарплат по городам', fontdict={'fontsize': 8})
        a_t.barh(list([str(a).replace(' ', '\n').replace('-', '-\n') for a in reversed(list(self.stats5.keys()))]), list(reversed(list(self.stats5.values()))), color='blue', height=0.5, align='center')
        a_t.yaxis.set_tick_params(labelsize=6)
        a_t.xaxis.set_tick_params(labelsize=8)
        a_t.grid(axis='x')

        a_four.set_title('Доля вакансий по городам', fontdict={'fontsize': 8})
        h = 1 - sum([value for value in self.stats6.values()])
        a_four.pie(list(self.stats6.values()) + [h], labels=list(self.stats6.keys()) + ['Другие'], textprops={'fontsize': 6})

        plt.tight_layout()
        plt.savefig('graph.png')

    def pdfs(self):
        env = Environment(loader=FileSystemLoader('../templates'))
        template = env.get_template("pdf_template.html")
        stats = []
        for y in self.stats1.keys():
            stats.append([y, self.stats1[y], self.stats2[y], self.stats3[y], self.stats4[y]])

        for k in self.stats6:
            self.stats6[k] = round(self.stats6[k] * 100, 2)

        pdf_template = template.render({'name': self.vacancy_name, 'path': '{0}/{1}'.format(pathlib.Path(__file__).parent.resolve(), 'graph.png'), 'stats': stats, 'stats5': self.stats5, 'stats6': self.stats6})

        pdfkit.from_string(pdf_template, 'report.pdf', options={"enable-local-file-access": ""})

    def exce(self):
        w_s1 = self.wb.active
        w_s1.title = 'Статистика по годам'
        w_s1.append(['Год', 'Средняя зарплата', 'Средняя зарплата - ' + self.vacancy_name, 'Количество вакансий', 'Количество вакансий - ' + self.vacancy_name])
        for y in self.stats1.keys():
            w_s1.append([y, self.stats1[y], self.stats3[y], self.stats2[y], self.stats4[y]])

        datas = [['Год ', 'Средняя зарплата ', ' Средняя зарплата - ' + self.vacancy_name, ' Количество вакансий', ' Количество вакансий - ' + self.vacancy_name]]
        cols_w = []
        for line in datas:
            for x, c in enumerate(line):
                if len(cols_w) > x:
                    if len(c) > cols_w[x]:
                        cols_w[x] = len(c)
                else:
                    cols_w += [len(c)]

        for x, column_width in enumerate(cols_w, 1):  # ,1 to start at 1
            w_s1.column_dimensions[get_column_letter(x)].width = column_width + 2

        datas = []
        datas.append(['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий'])
        for (c_f, v_f), (c_s, v_s) in zip(self.stats5.items(), self.stats6.items()):
            datas.append([c_f, v_f, '', c_s, v_s])
        w_s2 = self.wb.create_sheet('Статистика по городам')
        for line in datas:
            w_s2.append(line)

        cols_w = []
        for line in datas:
            for x, c in enumerate(line):
                c = str(c)
                if len(cols_w) > x:
                    if len(c) > cols_w[x]:
                        cols_w[x] = len(c)
                else:
                    cols_w += [len(c)]

        for x, column_width in enumerate(cols_w, 1):  # ,1 to start at 1
            w_s2.column_dimensions[get_column_letter(x)].width = column_width + 2

        font_bold = Font(bold=True)
        for c in 'ABCDE':
            w_s1[c + '1'].font = font_bold
            w_s2[c + '1'].font = font_bold

        for index, _ in enumerate(self.stats5):
            w_s2['E' + str(index + 2)].number_format = '0.00%'

        th = Side(border_style='thin',
                    color='00000000')

        for line in range(len(datas)):
            for c in 'ABDE':
                w_s2[c + str(line + 1)].border = Border(left=th,
                                                        bottom=th,
                                                        right=th,
                                                        top=th)

        for line, _ in enumerate(self.stats1):
            for c in 'ABCDE':
                w_s1[c + str(line + 1)].border = Border(left=th,
                                                        bottom=th,
                                                        right=th,
                                                        top=th)

        self.wb.save(filename='report.xlsx')


class Vacs:

    def __init__(self, vacs):
        self.name = vacs['name']
        self.salary_from = int(float(vacs['salary_from']))
        self.salary_to = int(float(vacs['salary_to']))
        self.salary_currency = vacs['salary_currency']
        self.salary_average = self.currency_to_rub[self.salary_currency] * (self.salary_from + self.salary_to) / 2
        self.area_name = vacs['area_name']
        self.year = int(vacs['published_at'][:4])

    currency_to_rub = {
        "EUR": 59.90,
        "KZT": 0.13,
        "RUR": 1,
        "GEL": 21.74,
        "UZS": 0.0055,
        "BYR": 23.91,
        "AZN": 35.68,
        "KGS": 0.76,
        "UAH": 1.64,
        "USD": 60.66,
    }

if __name__ == '__main__':
    Inputs()